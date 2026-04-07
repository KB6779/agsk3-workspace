"""
АГСК-3 Рабочее место — FastAPI сервер.
Запуск: python app.py
Поддержка: SQLite (локально) и Supabase PostgreSQL (продакшен).
"""

import io
import os
from copy import copy
from functools import lru_cache
from pathlib import Path
from datetime import date
from urllib.parse import quote

import httpx
import openpyxl
from openpyxl.styles import Font, Alignment
from fastapi import FastAPI, Query, Request
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware

# ═══════════════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════════════

SUPABASE_URL = os.getenv("SUPABASE_URL", "https://omykcphkzmmqpwswwfsw.supabase.co")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
if not SUPABASE_KEY:
    raise RuntimeError(
        "SUPABASE_KEY environment variable is required. "
        "Set it in Railway Variables (production) or .env (local)."
    )

TPL_FIRST = Path(__file__).parent / "tools" / "Шаблон" / "Шаблон Спецификации Первый лист.xlsx"
TPL_NEXT  = Path(__file__).parent / "tools" / "Шаблон" / "Шаблон Спецификации Последующие листы (2,3,....).xlsx"
STATIC_PATH = Path(__file__).parent / "static"

app = FastAPI(title="АГСК-3 Рабочее место")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# Supabase HTTP клиент
_http = httpx.Client(
    base_url=f"{SUPABASE_URL}/rest/v1",
    headers={"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"},
    timeout=15.0,
)


# ═══════════════════════════════════════════════════════════════════════════
# СПРАВОЧНИКИ
# ═══════════════════════════════════════════════════════════════════════════

SECTIONS = {
    "21": "Раздел 21 — Материалы строительные: нерудные, вяжущие, стеновые, кровельные, лесные, металлопрокат",
    "22": "Раздел 22 — Изделия: железобетонные, бетонные, столярные, металлические, кровельные",
    "23": "Раздел 23 — Материалы и изделия для фасадов, отделки, изоляции, кровли",
    "24": "Раздел 24 — Трубы, арматура, кабели, сантехника, отопление, вентиляция, электротехника, связь",
    "25": "Раздел 25 — Дорожные знаки, изоляторы, взрывчатые вещества, бортовой камень",
    "27": "Раздел 27 — Дренажные грунты, металлоконструкции, облицовка, дорожные стойки, озеленение",
    "28": "Раздел 28 — Озеленение и благоустройство",
    "51": "Раздел 51 — Насосное и котельное оборудование, вентиляторы, трансформаторы, подъёмные платформы",
    "52": "Раздел 52 — Оборудование: игровое, учебное, медицинское, досмотровое",
    "53": "Раздел 53 — Установки очистки воды",
    "54": "Раздел 54 — Насосные установки, кондиционирование, специальное оборудование",
    "55": "Раздел 55 — Трансформаторы, видеокамеры, кабели, окна, отопление, АСУТП, фиброцемент",
}
SECTIONS_LIST = [{"code": k, "name": v} for k, v in SECTIONS.items()]


# ═══════════════════════════════════════════════════════════════════════════
# ШАБЛОН ГОСТ 21.110 — константы
# ═══════════════════════════════════════════════════════════════════════════

ITEM_COLS = {"pos": 7, "name": 8, "type_mark": 9, "agsk_code": 10,
             "manufacturer": 14, "unit": 18, "qty": 19, "weight": 20, "note": 22}

DATA_ROWS_FIRST = list(range(4, 14)) + list(range(15, 30, 2))
DATA_ROWS_NEXT = list(range(4, 20)) + list(range(20, 32, 2))

STAMP_MAP = {
    "dev_name": (39, 12), "dev_date": (39, 16),
    "check_name": (40, 12), "check_date": (40, 16),
    "norm_name": (43, 12), "norm_date": (43, 16),
    "appr_name": (44, 12), "appr_date": (44, 16),
    "project": (39, 17), "stage": (40, 21),
    "sheet": (40, 23), "sheets": (40, 24), "system": (42, 21),
}
GROUP_ORDER = {"Оборудование": 0, "Изделия": 1, "Материалы": 2}

FIELDS = "id,code,name,standard,unit,price_estimated,price_release"


# ═══════════════════════════════════════════════════════════════════════════
# DATABASE — Supabase PostgreSQL
# ═══════════════════════════════════════════════════════════════════════════

def _sb_get(params: dict) -> list:
    """GET запрос к Supabase REST API."""
    r = _http.get("/agsk_catalog", params=params)
    r.raise_for_status()
    return r.json()


@lru_cache(maxsize=64)
def _groups_cached(section: str) -> list:
    rows = _sb_get({
        "select": "code,name",
        "code": f"like.{section}%",
        "code": f"like.{section}%-0100",
        "order": "code",
        "limit": "200",
    })
    # Supabase не поддерживает 2 фильтра на одно поле через params —
    # используем and filter
    r = _http.get("/agsk_catalog", params={
        "select": "code,name",
        "and": f"(code.like.{section}%,code.like.%-0100)",
        "order": "code",
        "limit": "200",
    })
    r.raise_for_status()
    rows = r.json()

    seen = set()
    result = []
    for row in rows:
        grp = row["code"][:7]
        if grp not in seen:
            seen.add(grp)
            result.append({"code": grp, "name": row["name"]})
    return result


def _fts_search(q: str, prefix: str, limit: int) -> list:
    """Полнотекстовый поиск: каждое слово через отдельный ILIKE фильтр."""
    words = [w.strip() for w in q.split() if w.strip()]
    if not words:
        return []

    # Строим PostgREST AND фильтр: каждое слово через ilike
    # and=(name.ilike.%слово1%,name.ilike.%слово2%)
    conditions = [f"name.ilike.%{w}%" for w in words]
    if prefix:
        conditions.append(f"code.like.{prefix}")

    params = {
        "select": FIELDS,
        "and": f"({','.join(conditions)})",
        "order": "code",
        "limit": str(limit),
    }
    try:
        return _sb_get(params)
    except httpx.HTTPStatusError:
        # Fallback: один ILIKE по первому слову
        params = {"select": FIELDS, "name": f"ilike.%{words[0]}%",
                  "order": "code", "limit": str(limit)}
        if prefix:
            params["code"] = f"like.{prefix}"
        return _sb_get(params)


# ═══════════════════════════════════════════════════════════════════════════
# API ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/sections")
def api_sections():
    return JSONResponse(SECTIONS_LIST, headers={"Cache-Control": "public, max-age=3600"})


@app.get("/api/groups")
def api_groups(section: str = ""):
    if not section:
        return []
    result = _groups_cached(section)
    return JSONResponse(result, headers={"Cache-Control": "public, max-age=3600"})


@app.get("/api/search")
def api_search(q: str = Query(""), section: str = "", group: str = "", limit: int = 50):
    q = q.strip()
    prefix = f"{group}%" if group else (f"{section}%" if section else "")

    if not q and not prefix:
        return _sb_get({"select": FIELDS, "order": "id", "limit": str(limit)})
    elif not q:
        return _sb_get({"select": FIELDS, "code": f"like.{prefix}", "order": "code", "limit": str(limit)})
    else:
        return _fts_search(q, prefix, limit)


@app.get("/api/match")
def api_match(name: str = "", section: str = "", limit: int = 5):
    return api_search(q=name, section=section, group="", limit=limit)


@app.get("/api/context")
def api_context(code: str = ""):
    """Контекст позиции: родительское наименование (ст.2) + характеристика (ст.3)."""
    if not code:
        return {"col2_name": "", "col3_type": "", "unit": ""}

    rows = _sb_get({"select": "name,standard,unit", "code": f"eq.{code}", "limit": "1"})
    if not rows:
        return {"col2_name": "", "col3_type": "", "unit": ""}

    item = rows[0]
    item_name = item["name"]
    item_std = item["standard"] or ""
    item_unit = item["unit"] or ""

    # Ищем ближайшего описательного родителя
    parent_name = ""
    segments = code.split("-")
    candidates = []
    if len(segments) == 4:
        candidates.append("-".join(segments[:3]))
    if len(segments) >= 3:
        base = "-".join(segments[:2])
        last = segments[2] if len(segments) >= 3 else segments[-1]
        if len(last) >= 4 and last != "0100":
            candidates.append(f"{base}-{last[:2]}00")
        if last != "0100":
            candidates.append(f"{base}-0100")

    for cand in candidates:
        if cand == code:
            continue
        rows2 = _sb_get({"select": "name", "code": f"eq.{cand}", "limit": "1"})
        if rows2 and rows2[0]["name"] and rows2[0]["name"] != item_name:
            name_lower = rows2[0]["name"].lower()
            if name_lower.startswith("продукция ") or name_lower.startswith("оборудование фирмы "):
                continue
            parent_name = rows2[0]["name"]
            break

    # Умное разделение
    SPLIT_KW = ["модели ", "модель ", "марки ", "марка ", "размерами ", "размером "]
    col2, col3 = "", ""
    starts_upper = item_name and item_name[0].isupper()

    if starts_upper:
        name_lower = item_name.lower()
        for kw in SPLIT_KW:
            pos = name_lower.find(kw)
            if pos > 5:
                col2 = item_name[:pos].strip().rstrip(",").strip()
                col3 = item_name[pos:].strip()
                break

    if col2:
        if item_std:
            col3 = f"{col3} {item_std}" if col3 else item_std
        return {"col2_name": col2, "col3_type": col3, "unit": item_unit}
    elif parent_name:
        col3 = item_name
        if item_std:
            col3 = f"{item_name} {item_std}"
        return {"col2_name": parent_name, "col3_type": col3, "unit": item_unit}
    else:
        return {"col2_name": item_name, "col3_type": item_std, "unit": item_unit}


# ═══════════════════════════════════════════════════════════════════════════
# ЭКСПОРТ ГОСТ 21.110
# ═══════════════════════════════════════════════════════════════════════════

def _fill_sheet_items(ws, items: list, data_rows: list,
                      pos_start: int = 1, group_state: str = "") -> tuple:
    current_group = group_state
    pos = pos_start
    slot = 0
    consumed = 0

    for item in items:
        group_type = item.get("group_type", "Оборудование")

        if group_type != current_group:
            if slot >= len(data_rows):
                break
            row = data_rows[slot]
            ws.cell(row=row, column=ITEM_COLS["name"]).value = group_type
            ws.cell(row=row, column=ITEM_COLS["name"]).font = Font(name="Calibri", size=11, bold=True)
            ws.cell(row=row, column=ITEM_COLS["name"]).alignment = Alignment(horizontal="left", vertical="center")
            slot += 1
            current_group = group_type

        if slot >= len(data_rows):
            break

        row = data_rows[slot]
        ws.cell(row=row, column=ITEM_COLS["pos"]).value = str(pos)
        for field in ("name", "type_mark", "agsk_code", "manufacturer", "unit", "qty", "weight", "note"):
            ws.cell(row=row, column=ITEM_COLS[field]).value = item.get(field) or ""
        slot += 1
        pos += 1
        consumed += 1

    return pos, consumed, current_group


def _fill_stamp(ws, stamp: dict) -> None:
    field_map = {
        "dev_name": "developer", "dev_date": "dev_date",
        "check_name": "checker", "check_date": "check_date",
        "norm_name": "norm_ctrl", "norm_date": "norm_date",
        "appr_name": "approver", "appr_date": "appr_date",
        "stage": "stage", "sheet": "sheet", "sheets": "sheets",
    }
    for stamp_key, payload_key in field_map.items():
        row, col = STAMP_MAP[stamp_key]
        ws.cell(row=row, column=col).value = stamp.get(payload_key, "")

    code = stamp.get("code", "")
    name = stamp.get("name", "")
    row, col = STAMP_MAP["project"]
    ws.cell(row=row, column=col).value = f"{code}\n{name}" if name else code

    system = stamp.get("system", "")
    if system:
        row, col = STAMP_MAP["system"]
        ws.cell(row=row, column=col).value = system


@app.post("/api/export/gost")
async def api_export_gost(request: Request):
    payload = await request.json()
    stamp = payload.get("stamp", {})
    items = payload.get("items", [])

    items_sorted = sorted(
        items, key=lambda x: GROUP_ORDER.get(x.get("group_type", "Оборудование"), 99),
    )
    remaining = list(items_sorted)

    wb = openpyxl.load_workbook(str(TPL_FIRST))
    ws1 = wb.worksheets[0]
    ws1.title = "Лист 1"

    next_pos, consumed, last_group = _fill_sheet_items(
        ws1, remaining, DATA_ROWS_FIRST, pos_start=1, group_state="",
    )
    remaining = remaining[consumed:]
    _fill_stamp(ws1, stamp)

    sheet_num = 2
    while remaining:
        tpl_next = openpyxl.load_workbook(str(TPL_NEXT))
        ws_src = tpl_next.worksheets[0]

        next_pos, consumed, last_group = _fill_sheet_items(
            ws_src, remaining, DATA_ROWS_NEXT, pos_start=next_pos, group_state=last_group,
        )
        remaining = remaining[consumed:]
        ws_src.cell(row=36, column=25).value = sheet_num

        new_ws = wb.create_sheet(f"Лист {sheet_num}")
        for col_letter, dim in ws_src.column_dimensions.items():
            if dim.width:
                new_ws.column_dimensions[col_letter].width = dim.width
        for row_num, dim in ws_src.row_dimensions.items():
            if dim.height:
                new_ws.row_dimensions[row_num].height = dim.height
        for mc in ws_src.merged_cells.ranges:
            new_ws.merge_cells(str(mc))
        for row in ws_src.iter_rows(min_row=1, max_row=ws_src.max_row, max_col=ws_src.max_column):
            for c in row:
                dst = new_ws.cell(row=c.row, column=c.column, value=c.value)
                if c.has_style:
                    dst.font = copy(c.font)
                    dst.alignment = copy(c.alignment)
                    dst.border = copy(c.border)
                    dst.fill = copy(c.fill)
                    dst.number_format = c.number_format
        new_ws.page_setup.orientation = "landscape"
        new_ws.page_setup.paperSize = 8
        sheet_num += 1

    total_sheets = sheet_num - 1
    ws1.cell(row=40, column=24).value = str(total_sheets)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    proj = stamp.get("code", "spec").replace("/", "-").replace("\\", "-")
    safe_name = f"Spec_{date.today().isoformat()}.xlsx"
    utf_name = f"Spec_{proj}_{date.today().isoformat()}.xlsx"
    cd = f"attachment; filename=\"{safe_name}\"; filename*=UTF-8''{quote(utf_name)}"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": cd},
    )


# ═══════════════════════════════════════════════════════════════════════════
# СТАТИКА + ЗАПУСК
# ═══════════════════════════════════════════════════════════════════════════

STATIC_PATH.mkdir(exist_ok=True)
app.mount("/", StaticFiles(directory=str(STATIC_PATH), html=True), name="static")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app:app", host="127.0.0.1", port=9000, reload=False)

"""
Конвертер АГСК-3 Excel → SQLite с FTS5 для быстрого поиска.
Запуск: python tools/convert_agsk.py
"""

import sys
import io
import sqlite3
import time
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

EXCEL_PATH = r"C:\Users\KAIRAT_BAIKULOV\Desktop\АГСК-3\АГСК-3_март 2026 all inclusive.xlsx"
DB_PATH = "tools/agsk.db"


def convert():
    print("Загружаю Excel (это займёт ~30 сек)...")
    t0 = time.time()
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    print(f"Загружено за {time.time()-t0:.1f}с. Строк: {ws.max_row}")

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    cur.executescript("""
        DROP TABLE IF EXISTS catalog;
        DROP TABLE IF EXISTS catalog_fts;

        CREATE TABLE catalog (
            id       INTEGER PRIMARY KEY,
            code     TEXT NOT NULL,
            name     TEXT NOT NULL,
            standard TEXT,
            unit     TEXT,
            price_estimated REAL,
            price_release   REAL
        );

        CREATE INDEX idx_code ON catalog(code);
    """)

    print("Записываю данные...")
    t1 = time.time()
    batch = []
    skipped = 0

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        code, name, standard, unit, price_est, price_rel = row

        if not code or not name:
            skipped += 1
            continue

        batch.append((
            str(code).strip(),
            str(name).strip(),
            str(standard).strip() if standard else None,
            str(unit).strip() if unit else None,
            float(price_est) if price_est is not None else None,
            float(price_rel) if price_rel is not None else None,
        ))

        if len(batch) >= 5000:
            cur.executemany(
                "INSERT INTO catalog (code,name,standard,unit,price_estimated,price_release) VALUES (?,?,?,?,?,?)",
                batch,
            )
            batch.clear()
            if i % 50000 == 0:
                print(f"  {i} строк обработано...")

    if batch:
        cur.executemany(
            "INSERT INTO catalog (code,name,standard,unit,price_estimated,price_release) VALUES (?,?,?,?,?,?)",
            batch,
        )

    print(f"Данные записаны за {time.time()-t1:.1f}с. Пропущено: {skipped}")

    print("Создаю FTS5-индекс (полнотекстовый поиск)...")
    t2 = time.time()
    cur.executescript("""
        CREATE VIRTUAL TABLE catalog_fts USING fts5(
            code,
            name,
            standard,
            content='catalog',
            content_rowid='id',
            tokenize='unicode61'
        );

        INSERT INTO catalog_fts(rowid, code, name, standard)
        SELECT id, code, name, COALESCE(standard,'') FROM catalog;
    """)
    print(f"FTS-индекс создан за {time.time()-t2:.1f}с")

    conn.commit()
    conn.close()

    total = time.time() - t0
    print(f"\nГотово! База: {DB_PATH}")
    print(f"Общее время: {total:.1f}с")


if __name__ == "__main__":
    convert()
